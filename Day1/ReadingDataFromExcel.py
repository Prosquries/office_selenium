import openpyxl
# Reading Opertaions
file = r"D:\Aarav\Selenium\Day1\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

# Print the data from the excel sheet

for row in range(1,rows+1):
    for col in range(1,cols+1):
        print(sheet.cell(row,col).value,end = '   ')
    print()

# Writing operations
File = r"D:\Aarav\Selenium\Day1\Multiple.xlsx"
Workbook = openpyxl.load_workbook(File)
Sheet = Workbook.active

## Same Data
# for r in range(1,6):
#     for c in range(1,4):
#         Sheet.cell(r,c).value = "Welcome"
#
# Workbook.save(File)

# Multiple Data
Sheet.cell(1,1).value="123"
Sheet.cell(1,2).value="Aarav"

Sheet.cell(2,1).value="456"
Sheet.cell(2,2).value="John"

Sheet.cell(3,1).value="789"
Sheet.cell(3,2).value="David"

Sheet.cell(4,1).value="147"
Sheet.cell(4,2).value="Manager"

Workbook.save(File)
